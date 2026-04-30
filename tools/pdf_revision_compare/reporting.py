"""Reviewer-focused report generation for IFB/GMP PDF compare runs."""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
import csv
import json
import traceback

try:  # pragma: no cover - optional import in minimal test environments
    from openpyxl import Workbook as OpenPyxlWorkbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:  # noqa: BLE001
    OpenPyxlWorkbook = None  # type: ignore[assignment]
    Alignment = Font = PatternFill = get_column_letter = None  # type: ignore[assignment]

try:  # pragma: no cover - fallback writer
    from .xlsxlite import SheetSpec, write_xlsx
except ImportError:  # pragma: no cover
    from xlsxlite import SheetSpec, write_xlsx  # type: ignore

__all__ = [
    "build_report_model",
    "build_report_sheets",
    "build_csv_manifest_rows",
    "normalize_records",
    "normalize_value",
    "write_manifest_csv",
    "write_report_bundle",
]


SECTION_ORDER = [
    "Run Summary",
    "Sheet Map",
    "Changed Sheets",
    "Schedule Row Diffs",
    "Title Block Diffs",
    "Visual Page Diffs",
    "Added Sheets",
    "Removed Sheets",
    "Review Needed",
    "Exceptions",
]

SECTION_ALIASES = {
    "Run Summary": ("run_summary", "summary", "runSummary", "Summary"),
    "Sheet Map": ("sheet_map", "document_summary", "pairing_index", "documents", "pairings"),
    "Changed Sheets": ("changed_sheets",),
    "Schedule Row Diffs": ("schedule_row_diffs", "structured_row_diffs", "row_diffs", "rows"),
    "Title Block Diffs": ("title_block_diffs",),
    "Visual Page Diffs": ("visual_page_diffs", "page_diffs", "pages"),
    "Added Sheets": ("added_sheets",),
    "Removed Sheets": ("removed_sheets",),
    "Review Needed": ("review_needed",),
    "Exceptions": ("exceptions", "errors", "warnings", "issues"),
}

SECTION_COLUMNS = {
    "Run Summary": ["metric", "value"],
    "Sheet Map": [
        "sheet_id",
        "sheet_title",
        "family",
        "status",
        "compare_status",
        "revision_status",
        "scope_changed",
        "ifb_name",
        "gmp_name",
        "ifb_revision",
        "gmp_revision",
        "changed_pages",
        "highlighted_before_pdf",
        "highlighted_after_pdf",
        "packaged_path",
    ],
    "Changed Sheets": [
        "sheet_id",
        "sheet_title",
        "family",
        "compare_status",
        "revision_status",
        "scope_changed",
        "changed_pages",
        "schedule_row_diff_count",
        "title_block_diff_count",
        "visual_diff_count",
        "highlighted_before_pdf",
        "highlighted_after_pdf",
        "packaged_path",
    ],
    "Schedule Row Diffs": [
        "sheet_id",
        "sheet_title",
        "page_index",
        "region_id",
        "change_type",
        "before_text",
        "after_text",
        "before_cells",
        "after_cells",
        "severity",
        "confidence",
        "is_scope_diff",
        "before_changed_bboxes",
        "after_changed_bboxes",
        "highlighted_before_pdf",
        "highlighted_after_pdf",
    ],
    "Title Block Diffs": [
        "sheet_id",
        "sheet_title",
        "page_label",
        "region_kind",
        "change_type",
        "before_text",
        "after_text",
        "highlighted_before_pdf",
        "highlighted_after_pdf",
    ],
    "Visual Page Diffs": [
        "sheet_id",
        "sheet_title",
        "family",
        "page_label",
        "change_type",
        "global_box_count",
        "emphasized_box_count",
        "highlighted_before_pdf",
        "highlighted_after_pdf",
        "notes",
    ],
    "Added Sheets": ["sheet_id", "sheet_title", "family", "gmp_name", "gmp_revision", "packaged_path"],
    "Removed Sheets": ["sheet_id", "sheet_title", "family", "ifb_name", "ifb_revision", "packaged_path"],
    "Review Needed": ["sheet_id", "sheet_title", "family", "status", "compare_status", "reasons", "packaged_path"],
    "Exceptions": ["scope", "sheet_id", "pair_id", "message"],
}


def normalize_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, datetime):
        return value.astimezone(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, Mapping):
        return {str(key): normalize_value(item) for key, item in value.items()}
    if isinstance(value, Sequence) and not isinstance(value, (str, bytes, bytearray)):
        return [normalize_value(item) for item in value]
    if isinstance(value, BaseException):
        payload = {"type": type(value).__name__, "message": str(value)}
        if value.__traceback__ is not None:
            payload["traceback"] = "".join(traceback.format_exception(type(value), value, value.__traceback__))
        return payload
    return str(value)


def _stringify(value: Any) -> Any:
    normalized = normalize_value(value)
    if normalized is None:
        return None
    if isinstance(normalized, (bool, int, float)):
        return normalized
    if isinstance(normalized, (dict, list)):
        return json.dumps(normalized, ensure_ascii=False, sort_keys=True)
    return str(normalized)


def normalize_records(records: Any) -> list[dict[str, Any]]:
    if records is None:
        return []
    if isinstance(records, Mapping):
        return [{str(key): normalize_value(value) for key, value in records.items()}]
    if isinstance(records, Sequence) and not isinstance(records, (str, bytes, bytearray)):
        normalized: list[dict[str, Any]] = []
        for index, item in enumerate(records):
            if isinstance(item, Mapping):
                normalized.append({str(key): normalize_value(value) for key, value in item.items()})
            else:
                normalized.append({"index": index, "value": normalize_value(item)})
        return normalized
    return [{"value": normalize_value(records)}]


def _extract_section(source: Any, aliases: Sequence[str]) -> Any:
    if source is None:
        return None
    if isinstance(source, Mapping):
        for key in aliases:
            if key in source:
                return source[key]
    for key in aliases:
        if hasattr(source, key):
            return getattr(source, key)
    return None


def _first_section(compare_results: Any, name: str) -> list[dict[str, Any]]:
    return normalize_records(_extract_section(compare_results, SECTION_ALIASES[name]))


def _infer_summary(compare_results: Any, sections: dict[str, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    explicit = _first_section(compare_results, "Run Summary")
    if explicit:
        return explicit
    sheet_rows = sections.get("Sheet Map", [])
    return [
        {"metric": "matched_pairs", "value": sum(1 for row in sheet_rows if row.get("status") == "changed" or row.get("status") == "unchanged")},
        {"metric": "changed_documents", "value": len(sections.get("Changed Sheets", []))},
        {"metric": "added_documents", "value": len(sections.get("Added Sheets", []))},
        {"metric": "removed_documents", "value": len(sections.get("Removed Sheets", []))},
        {"metric": "review_needed", "value": len(sections.get("Review Needed", []))},
        {"metric": "schedule_row_diffs", "value": len(sections.get("Schedule Row Diffs", []))},
        {"metric": "exceptions", "value": len(sections.get("Exceptions", []))},
    ]


def _derive_sections(compare_results: Any, manifests: Any, exceptions: Any) -> dict[str, list[dict[str, Any]]]:
    document_rows = _first_section(compare_results, "Sheet Map")
    pairing_rows = normalize_records(_extract_section(compare_results, ("pairing_index", "pairings")))
    page_rows = _first_section(compare_results, "Visual Page Diffs")
    region_rows = normalize_records(_extract_section(compare_results, ("region_diffs", "regions")))
    row_rows = _first_section(compare_results, "Schedule Row Diffs")
    exception_rows = _first_section(compare_results, "Exceptions") or normalize_records(exceptions)

    sheet_map = document_rows or pairing_rows or normalize_records(manifests)
    changed = [
        row for row in document_rows
        if row.get("changed") is True
        or row.get("scope_changed") is True
        or row.get("compare_status") in {"schedule_bom_changed", "drawing_scope_changed", "scope_changed"}
    ]
    added = [row for row in document_rows if row.get("status") == "added"]
    removed = [row for row in document_rows if row.get("status") == "removed"]
    review = [
        row for row in [*pairing_rows, *document_rows]
        if row.get("status") == "review_needed" or row.get("compare_status") == "compare_failed"
    ]
    schedule_rows = [
        row for row in row_rows
        if row.get("is_scope_diff") is True
    ]
    title_rows = [
        row for row in region_rows
        if row.get("region_kind") in {"metadata-like", "title-block-like"}
    ]
    visual_rows = [
        row for row in page_rows
        if row.get("change_type") != "unchanged" or int(row.get("global_box_count") or 0) > 0
    ]

    sections = {
        "Sheet Map": sheet_map,
        "Changed Sheets": changed,
        "Schedule Row Diffs": schedule_rows,
        "Title Block Diffs": title_rows,
        "Visual Page Diffs": visual_rows,
        "Added Sheets": added,
        "Removed Sheets": removed,
        "Review Needed": review,
        "Exceptions": exception_rows,
    }
    sections["Run Summary"] = _infer_summary(compare_results, sections)
    return sections


def build_report_model(compare_results: Any = None, manifests: Any = None, exceptions: Any = None) -> dict[str, Any]:
    sections = _derive_sections(compare_results, manifests, exceptions)
    source_manifests = normalize_records(manifests)
    return {
        "generated_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "sections": sections,
        "counts": {name: len(rows) for name, rows in sections.items()},
        "source_manifests": source_manifests,
        "source_manifest_count": len(source_manifests),
    }


def _section_headers(name: str, rows: list[dict[str, Any]]) -> list[str]:
    preferred = list(SECTION_COLUMNS.get(name, []))
    if preferred:
        return preferred
    seen = set(preferred)
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                preferred.append(key)
    return preferred or ["message"]


def build_report_sheets(model: dict[str, Any]) -> list[SheetSpec]:
    sections: dict[str, list[dict[str, Any]]] = model.get("sections", {})
    return [
        SheetSpec(name=name, rows=sections.get(name, []), columns=_section_headers(name, sections.get(name, [])))
        for name in SECTION_ORDER
    ]


def build_csv_manifest_rows(model: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for section_name in SECTION_ORDER:
        for index, row in enumerate(model.get("sections", {}).get(section_name, [])):
            flat_row = {"section": section_name, "row_index": index}
            for key, value in row.items():
                flat_row[key] = _stringify(value)
            rows.append(flat_row)
    for index, row in enumerate(model.get("source_manifests", [])):
        flat_row = {"section": "Source Manifest", "row_index": index}
        for key, value in row.items():
            flat_row[key] = _stringify(value)
        rows.append(flat_row)
    return rows


def write_manifest_csv(path: str | Path, model: dict[str, Any]) -> Path:
    out_path = Path(path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    rows = build_csv_manifest_rows(model)
    headers: list[str] = ["section", "row_index"]
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    with out_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: _stringify(value) for key, value in row.items()})
    return out_path


def _is_path_column(header: str) -> bool:
    lowered = header.lower()
    return lowered.endswith("_path") or "pdf" in lowered or "manifest" in lowered or "folder" in lowered


def _write_openpyxl(path: Path, sheets: list[SheetSpec]) -> None:
    if OpenPyxlWorkbook is None:
        raise RuntimeError("openpyxl is not available")
    wb = OpenPyxlWorkbook()
    default = wb.active
    wb.remove(default)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    link_font = Font(color="0563C1", underline="single")
    status_fills = {
        "changed": PatternFill("solid", fgColor="F4CCCC"),      # Red (Modified)
        "added": PatternFill("solid", fgColor="77ECF5"),        # Aqua Blue (New)
        "removed": PatternFill("solid", fgColor="C6EFCE"),      # Green (Deleted)
        "moved": PatternFill("solid", fgColor="FCE4EC"),        # Pink (Moved only)
        "review_needed": PatternFill("solid", fgColor="FCE4EC"), # Pink (Manual Review)
        "compare_failed": PatternFill("solid", fgColor="F4CCCC"),
    }
    for spec in sheets:
        ws = wb.create_sheet(spec.name[:31])
        columns = list(spec.columns)
        ws.append(columns)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row in spec.rows:
            values = [_stringify(row.get(column)) for column in columns]
            ws.append(values)
            current_row = ws.max_row
            status = str(row.get("status") or row.get("compare_status") or "").lower()
            fill = status_fills.get(status)
            for col_index, column in enumerate(columns, start=1):
                cell = ws.cell(current_row, col_index)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if fill is not None:
                    cell.fill = fill
                if _is_path_column(column) and cell.value:
                    cell.hyperlink = str(cell.value)
                    cell.font = link_font
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col_index, column in enumerate(columns, start=1):
            letter = get_column_letter(col_index)
            max_width = max([len(str(column)), *[len(str(ws.cell(row=row, column=col_index).value or "")) for row in range(2, min(ws.max_row, 80) + 1)]])
            ws.column_dimensions[letter].width = min(max(max_width + 2, 12), 70)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_report_bundle(
    output_dir: str | Path,
    compare_results: Any = None,
    manifests: Any = None,
    exceptions: Any = None,
    *,
    workbook_name: str = "report.xlsx",
    manifest_json_name: str = "manifest.json",
    manifest_csv_name: str = "manifest.csv",
    write_manifest_csv_file: bool = True,
) -> dict[str, Any]:
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    model = build_report_model(compare_results=compare_results, manifests=manifests, exceptions=exceptions)
    sheets = build_report_sheets(model)
    workbook_path = out_dir / workbook_name
    try:
        _write_openpyxl(workbook_path, sheets)
        writer = "openpyxl"
        workbook_warning = None
    except Exception as exc:  # noqa: BLE001
        write_xlsx(workbook_path, sheets, title="IFB GMP PDF Compare Report", creator="PDFCompare")
        writer = "xlsxlite"
        workbook_warning = f"openpyxl failed; used xlsxlite fallback: {type(exc).__name__}: {exc}"

    manifest_path = out_dir / manifest_json_name
    manifest_payload = {
        **model,
        "workbook": workbook_path.name,
        "workbook_path": str(workbook_path),
        "workbook_writer": writer,
        "workbook_warning": workbook_warning,
        "manifest_csv": manifest_csv_name if write_manifest_csv_file else None,
    }
    manifest_path.write_text(json.dumps(manifest_payload, indent=2, ensure_ascii=False), encoding="utf-8")

    manifest_csv_path = None
    if write_manifest_csv_file:
        manifest_csv_path = write_manifest_csv(out_dir / manifest_csv_name, model)

    return {
        "output_dir": str(out_dir),
        "workbook_path": str(workbook_path),
        "manifest_json_path": str(manifest_path),
        "manifest_csv_path": str(manifest_csv_path) if manifest_csv_path else None,
        "model": model,
        "csv_rows": build_csv_manifest_rows(model),
        "sheets": [sheet.name for sheet in sheets],
        "workbook_writer": writer,
    }
