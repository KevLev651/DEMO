"""Comparator: match an Excel tag list against a P&ID PDF.

Usage (CLI):
    python comparator/comparator.py <pdf> <xlsx> [--mode instrument|equipment|line|all]

Outputs:
  - <xlsx>_annotated.xlsx : your Excel with added Comparison columns, plus
    new sheets: Comparison Summary, Comparison Results,
    PDF Candidates Not In Excel, PDF Page Diagnostics.
  - <pdf>_annotated.pdf   : highlights matched tags in yellow on the PDF.
"""

import argparse
import os
import re
import sys

import fitz
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


TOOL_DIR = os.path.abspath(os.path.dirname(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(TOOL_DIR, os.pardir))
for path in (TOOL_DIR, PROJECT_ROOT):
    if path not in sys.path:
        sys.path.insert(0, path)

from .pid_common import (
    validate_input_file,
    validate_pdf_content,
)
from .pid_pdf_common import (
    Tag,
    build_sheet_map,
    confidence_counts,
    confidence_rating,
    dedupe,
    normalize_sheet_id,
    parse_excel_tag,
    scan_open_document,
    summarize_profiles,
    tag_sort_key,
)


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HIGH_FILL = PatternFill("solid", fgColor="FFF2CC")   # light yellow (High)
MED_FILL = PatternFill("solid", fgColor="F4CCCC")    # light red (Everything Else)
LOW_FILL = PatternFill("solid", fgColor="F4CCCC")    # light red (Everything Else)

# Status-based fills: Yellow = High Match, Red = Everything Else
MATCH_FILL = PatternFill("solid", fgColor="F4CCCC")   # light red
MISS_FILL  = PatternFill("solid", fgColor="F4CCCC")   # light red
YELLOW_HIGHLIGHT = [1, 1, 0]
RED_HIGHLIGHT = [1, 0.8, 0.8]  # Light red highlight
CONFIDENCE_RANK = {"All": 0, "Low": 1, "Medium": 2, "High": 3}

TAG_HEADERS = [
    "instrument tag", "instrument id", "tag number", "tag no",
    "component tag", "item label", "phase 2a tag", "phase 2 tag", "tag",
    "full tag",
]
SHEET_HEADERS = ["p&id", "pid", "drawing", "dwg", "sheet"]


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def validate_xlsx(xlsx_path):
    try:
        wb = load_workbook(xlsx_path, read_only=True)
        wb.close()
    except Exception as exc:
        raise ValueError("Not a valid .xlsx file") from exc


def _style_header(row):
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _format_sheet(sheet, widths):
    sheet.freeze_panes = "A2"
    if sheet.max_row > 1:
        sheet.auto_filter.ref = sheet.dimensions
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width


def _rating_fill(rating):
    return HIGH_FILL if rating == "High" else MED_FILL


def _status_fill(status, rating=""):
    """Fill for user-facing Excel tag rows/cells."""
    if status and str(status).startswith("Matched"):
        return HIGH_FILL if rating == "High" else MATCH_FILL
    return MISS_FILL


def _save(wb, preferred):
    try:
        wb.save(preferred)
        return preferred
    except PermissionError:
        base, ext = os.path.splitext(preferred)
        for i in range(1, 100):
            fb = f"{base}_{i}{ext}"
            try:
                wb.save(fb)
                return fb
            except PermissionError:
                continue
        raise


def _score_header(header, keywords):
    """Score a header row by presence of tag/sheet column keywords."""
    best_col = None
    best_score = 0
    for idx, value in enumerate(header):
        s = str(value or "").strip().lower().replace("\n", " ")
        for rank, kw in enumerate(keywords):
            priority = len(keywords) - rank
            if kw == "tag" and s != "tag":
                continue
            if s == kw:
                score = priority * 100
            elif s.startswith(kw):
                score = priority * 10
            elif kw in s:
                score = priority
            else:
                continue
            if score > best_score:
                best_col = idx + 1
                best_score = score
    return best_col


def _find_header(sheet):
    """Find the header row, tag column, and optional P&ID/sheet column."""
    best = None
    for row_num in range(1, min(sheet.max_row, 15) + 1):
        header = [cell.value for cell in sheet[row_num]]
        if len([v for v in header if v not in (None, "")]) < 2:
            continue
        tag_col = _score_header(header, TAG_HEADERS)
        sheet_col = _score_header(header, SHEET_HEADERS)
        if tag_col:
            score = 10 + (5 if sheet_col else 0)
            if not best or score > best[0]:
                best = (score, row_num, tag_col, sheet_col)
    if not best:
        return None, None, None
    return best[1], best[2], best[3]


def _append_status_cols(sheet, header_row):
    start = sheet.max_column + 1
    labels = ["Comparison Status", "Best Confidence", "Rating",
              "PDF Occurrences", "Best PDF Sheet", "Best PDF Page",
              "Best PDF Zone", "Best PDF Method", "Best Evidence",
              "Best Text Source", "Best Source Agreement", "Best OCR Confidence",
              "Allowed Pages", "Sheet Filter Used", "Filter Reason",
              "Source Node IDs", "Source Node Count"]
    for off, label in enumerate(labels):
        cell = sheet.cell(header_row, start + off, label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    return {
        "status": start, "confidence": start + 1, "rating": start + 2,
        "count": start + 3, "sheet": start + 4, "page": start + 5,
        "zone": start + 6, "method": start + 7, "evidence": start + 8,
        "text_source": start + 9, "source_agreement": start + 10,
        "ocr_confidence": start + 11, "allowed_pages": start + 12,
        "sheet_filter": start + 13, "filter_reason": start + 14,
        "source_node_ids": start + 15, "source_node_count": start + 16,
    }


# ---------------------------------------------------------------------------
# Matching
# ---------------------------------------------------------------------------

def _compact_match_text(value):
    """Format-insensitive comparison key for block fallback matching."""
    return re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())


def _classify_block_zone(display_rect, page_rect):
    w = max(page_rect.width, 1.0)
    h = max(page_rect.height, 1.0)
    cx = (display_rect.x0 + display_rect.x1) / 2
    cy = (display_rect.y0 + display_rect.y1) / 2
    if cx > w * 0.84:
        return "title-block/right"
    if cy > h * 0.90:
        return "title-block/bottom"
    if cy < h * 0.04 or cx < w * 0.02 or cx > w * 0.98:
        return "edge"
    return "drawing"


def _build_block_fallback_index(doc, page_to_sheet):
    """Build page text blocks in display coordinates for PIDTool-style fallback.

    This is intentionally pattern-light: it does not infer tags from the PDF.
    It only lets known Excel tags match if their parsed root/component text is
    contained in a nearby PDF text block.
    """
    out = {}
    for page_num, page in enumerate(doc, start=1):
        sheet_id = page_to_sheet.get(page_num, f"PAGE_{page_num}")
        blocks = []
        for block_idx, block in enumerate(page.get_text("blocks")):
            if len(block) < 5:
                continue
            text = str(block[4] or "").strip()
            if not text:
                continue
            display = fitz.Rect(block[0], block[1], block[2], block[3]) * page.rotation_matrix
            display.normalize()
            blocks.append({
                "page_num": page_num,
                "sheet_id": sheet_id,
                "block_id": f"p{page_num}:block:{block_idx}",
                "x0": display.x0,
                "y0": display.y0,
                "x1": display.x1,
                "y1": display.y1,
                "zone": _classify_block_zone(display, page.rect),
                "text": text,
                "compact": _compact_match_text(text),
            })
        out[page_num] = blocks
    return out


def _fallback_block_hits(parsed, allowed, block_index):
    if not parsed or not block_index:
        return []
    pages = sorted(allowed) if allowed else sorted(block_index)
    root_key = _compact_match_text(parsed.get("root", ""))
    comp_key = _compact_match_text(parsed.get("component", ""))
    full_key = _compact_match_text(parsed.get("full_tag", ""))
    tag_class = parsed.get("tag_class", "")

    hits = []
    for page_num in pages:
        for block in block_index.get(page_num, []):
            compact = block["compact"]
            matched = False
            if tag_class == "instrument" and root_key and comp_key:
                matched = (root_key in compact and comp_key in compact) or (full_key and full_key in compact)
            elif tag_class == "equipment" and root_key:
                matched = root_key in compact
            elif tag_class == "line":
                matched = (full_key and full_key in compact) or (root_key and comp_key and root_key in compact and comp_key in compact)
            if matched:
                hits.append(block)

    # Cross-block fallback for instruments: root in one block, component in
    # another block ON THE SAME PAGE and within spatial proximity. Catches
    # tags whose PDF text runs don't merge into a single text block.
    if not hits and tag_class == "instrument" and root_key and comp_key:
        for page_num in pages:
            blocks_on_page = block_index.get(page_num, [])
            if not blocks_on_page:
                continue
            root_blocks = [b for b in blocks_on_page if root_key in b["compact"]]
            comp_blocks = [b for b in blocks_on_page if comp_key in b["compact"]]
            if not root_blocks or not comp_blocks:
                continue
            for rb in root_blocks:
                rcx = (rb["x0"] + rb["x1"]) / 2
                rcy = (rb["y0"] + rb["y1"]) / 2
                for cb in comp_blocks:
                    if cb is rb:
                        continue
                    ccx = (cb["x0"] + cb["x1"]) / 2
                    ccy = (cb["y0"] + cb["y1"]) / 2
                    if abs(rcx - ccx) < 400 and abs(rcy - ccy) < 400:
                        # Represent the match at the root block location but
                        # tag it so confidence scoring can de-rate it.
                        merged = dict(rb)
                        merged["text"] = f'{rb["text"]} / {cb["text"]}'
                        merged["compact"] = rb["compact"] + cb["compact"]
                        merged["_cross_block"] = True
                        hits.append(merged)
                        break
                if hits:
                    break
            if hits:
                break
    return hits


def _fallback_tag(parsed, block, n_hits):
    confidence = 0.76
    if block["zone"] != "drawing":
        confidence -= 0.12
    if n_hits > 1:
        confidence -= min(0.12, 0.025 * (n_hits - 1))
    confidence = max(0.50, min(0.84, confidence))
    rating = confidence_rating(confidence)
    tag = Tag(
        full_tag=parsed["full_tag"],
        root=parsed["root"],
        component=parsed["component"],
        tag_class=parsed["tag_class"],
        method="block:fallback",
        source_spans=(),
        x0=block["x0"],
        y0=block["y0"],
        x1=block["x1"],
        y1=block["y1"],
        source_text=block["text"][:300],
        confidence=confidence,
        rating=rating,
        zone=block["zone"],
        evidence="block text fallback; parsed Excel tag contained in one PDF text block",
        sheet_id=block["sheet_id"],
        page_num=block["page_num"],
        text_source="native",
        source_agreement="native-only",
        page_ocr_mode="native",
        source_node_ids=block["block_id"],
        source_node_count=1,
    )
    return tag

def _index_tags(tags):
    by_full = {}
    by_root_comp = {}
    # Secondary indexes: tags with just the root (equipment, standalone:asset)
    # and tags with just the component (standalone:isa, or any instrument tag).
    # Used by the same-page fallback matcher when strict lookups miss.
    by_root_only = {}        # root -> list of tags having that root (any comp)
    by_comp_only = {}        # component -> list of instrument tags with that comp
    for t in tags:
        by_full.setdefault(t.full_tag, []).append(t)
        by_root_comp.setdefault((t.root, t.component), []).append(t)
        if t.root:
            by_root_only.setdefault(t.root, []).append(t)
        if t.component and t.tag_class == "instrument":
            by_comp_only.setdefault(t.component, []).append(t)
    return by_full, by_root_comp, by_root_only, by_comp_only


def _allowed_pages(sheet_value, sheet_map):
    if not sheet_value:
        return None  # search all pages
    sid = normalize_sheet_id(sheet_value)
    pages = sheet_map.get(sid)
    if not pages and sid.isdigit():
        matches = [v for k, v in sheet_map.items() if k.endswith(sid)]
        if len(matches) == 1:
            pages = matches[0]
    if not pages:
        return set()  # mapped/requested sheet is missing
    if isinstance(pages, (list, tuple, set)):
        return set(pages)
    return {pages}


def _allowed_page_details(sheet_value, sheet_map):
    allowed = _allowed_pages(sheet_value, sheet_map)
    if allowed is None:
        return None, "all pages", "No", "no sheet column/value supplied"
    if allowed == set():
        return allowed, "", "Yes", "sheet value not mapped to PDF"
    pages = sorted(allowed)
    reason = "duplicate sheet id" if len(pages) > 1 else "mapped sheet id"
    return allowed, ", ".join(str(p) for p in pages), "Yes", reason


def _choose_matches(parsed, allowed, by_full, by_root_comp,
                    by_root_only=None, by_comp_only=None):
    """Pick PDF-side matches for a parsed Excel tag.

    Tiered matching:
    1. Exact full_tag match
    2. (root, component) match
    3. Same-page root+component match: find a page where BOTH a tag with
       matching root AND a tag with matching component (any pairing state)
       appear. Synthesize a match at the component tag's location. This
       closes the "root exists, component missing" miss case where the
       scanner detected both pieces but could not geometrically pair them.
    """
    matches = list(by_full.get(parsed["full_tag"], []))
    if not matches:
        matches = list(by_root_comp.get((parsed["root"], parsed["component"]), []))
    if allowed:
        matches = [m for m in matches if m.page_num in allowed]
    if matches or not (by_root_only and by_comp_only):
        return sorted(matches, key=tag_sort_key)

    # Tier 3: same-page root+component fallback
    root = parsed.get("root", "")
    comp = parsed.get("component", "")
    if not root or not comp:
        return []
    root_tags = by_root_only.get(root, [])
    comp_tags = by_comp_only.get(comp, [])
    if not root_tags or not comp_tags:
        return []
    if allowed:
        root_tags = [t for t in root_tags if t.page_num in allowed]
        comp_tags = [t for t in comp_tags if t.page_num in allowed]
    if not root_tags or not comp_tags:
        return []

    # For each (root_tag, comp_tag) on same page, synthesize an inferred match.
    same_page_pairs = []
    root_pages = {t.page_num: t for t in root_tags}
    for ct in comp_tags:
        rt = root_pages.get(ct.page_num)
        if rt is None:
            continue
        same_page_pairs.append((rt, ct))
    if not same_page_pairs:
        return []

    synthesized = []
    for rt, ct in same_page_pairs:
        # Prefer the component tag's location as the match site.
        conf = 0.72
        # De-rate weak standalone:isa pairings slightly.
        if ct.method == "standalone:isa":
            conf = 0.66
        synth = Tag(
            full_tag=parsed["full_tag"],
            root=root,
            component=comp,
            tag_class=parsed.get("tag_class", "instrument"),
            method="samepage:root+comp",
            source_spans=tuple(list(rt.source_spans) + list(ct.source_spans)),
            x0=ct.x0, y0=ct.y0, x1=ct.x1, y1=ct.y1,
            source_text=f"{rt.source_text} + {ct.source_text}",
            confidence=conf,
            rating=confidence_rating(conf),
            zone=ct.zone,
            evidence=(
                "same-page inference: root and component detected separately "
                "on the same PDF page; pairing synthesized."
            ),
            sheet_id=ct.sheet_id,
            page_num=ct.page_num,
            text_source=ct.text_source,
            source_agreement=ct.source_agreement,
            page_ocr_mode=ct.page_ocr_mode,
            ocr_backend=ct.ocr_backend,
            ocr_reason=ct.ocr_reason,
            source_node_ids=ct.source_node_ids,
            source_node_count=ct.source_node_count,
            source_rects=ct.source_rects or ((ct.x0, ct.y0, ct.x1, ct.y1),),
        )
        synthesized.append(synth)
    return sorted(synthesized, key=tag_sort_key)


def _parsed_allowed_by_mode(parsed, mode):
    return mode == "all" or parsed.get("tag_class") == mode


def _match_confidence(candidate, allowed, n_matches):
    """Adjust per-match confidence for Excel sheet agreement and ambiguity."""
    conf = candidate.confidence
    notes = []
    if allowed and candidate.page_num in allowed:
        conf = min(0.99, conf + 0.05)
        notes.append("excel sheet match")
    if n_matches > 1:
        conf = max(0.05, conf - min(0.15, 0.04 * (n_matches - 1)))
        notes.append(f"ambiguous x{n_matches}")
    return conf, confidence_rating(conf), "; ".join(notes)


# ---------------------------------------------------------------------------
# Excel processing
# ---------------------------------------------------------------------------

def process_excel(xlsx_path, sheet_map, tags, block_index=None, mode="all"):
    tags = dedupe(tags, by_location=True)
    by_full, by_root_comp, by_root_only, by_comp_only = _index_tags(tags)
    wb = load_workbook(xlsx_path)

    stats = {"reviewed": 0, "matched": 0, "H": 0, "M": 0, "L": 0,
             "not_found": 0, "no_page": 0, "skipped": 0,
             "all_pages": 0, "ambiguous": 0, "fallback": 0,
             "samepage": 0}
    matched_tags = []
    matched_keys = set()
    result_rows = []

    found_a_header = False
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        if ws_name.lower() in {
            "summary", "comparison summary", "comparison results",
            "pdf candidates not in excel", "pdf page diagnostics", "page_index",
        }:
            continue
        hdr_row, tag_col, sheet_col = _find_header(ws)
        if not tag_col:
            continue
        found_a_header = True
        cols = _append_status_cols(ws, hdr_row)

        for row in ws.iter_rows(min_row=hdr_row + 1):
            if tag_col > len(row):
                continue
            tag_cell = row[tag_col - 1]
            raw = str(tag_cell.value).strip() if tag_cell.value is not None else ""
            if not raw:
                continue
            stats["reviewed"] += 1

            raw_sheet = ""
            if sheet_col and sheet_col <= len(row) and row[sheet_col - 1].value is not None:
                raw_sheet = str(row[sheet_col - 1].value).strip()

            allowed, allowed_label, sheet_filter_used, filter_reason = _allowed_page_details(raw_sheet, sheet_map)
            parsed = parse_excel_tag(raw)
            if not parsed:
                ws.cell(tag_cell.row, cols["status"], "Skipped - unrecognized tag")
                ws.cell(tag_cell.row, cols["rating"], "Low")
                ws.cell(tag_cell.row, cols["allowed_pages"], allowed_label)
                ws.cell(tag_cell.row, cols["sheet_filter"], sheet_filter_used)
                ws.cell(tag_cell.row, cols["filter_reason"], "unrecognized tag")
                tag_cell.fill = MISS_FILL
                stats["skipped"] += 1
                result_rows.append([ws_name, tag_cell.row, raw, raw_sheet,
                                    "Skipped - unrecognized tag", "", "", "Low",
                                    0, "", "", "", "", "", "", "", "",
                                    allowed_label, sheet_filter_used, "unrecognized tag", "", ""])
                continue

            if not _parsed_allowed_by_mode(parsed, mode):
                status = f"Skipped - outside {mode} mode"
                ws.cell(tag_cell.row, cols["status"], status)
                ws.cell(tag_cell.row, cols["rating"], "Low")
                ws.cell(tag_cell.row, cols["allowed_pages"], allowed_label)
                ws.cell(tag_cell.row, cols["sheet_filter"], sheet_filter_used)
                ws.cell(tag_cell.row, cols["filter_reason"], status)
                tag_cell.fill = MISS_FILL
                stats["skipped"] += 1
                result_rows.append([ws_name, tag_cell.row, raw, raw_sheet,
                                    status, parsed["full_tag"], "", "Low",
                                    0, "", "", "", "", "", "", "", "",
                                    allowed_label, sheet_filter_used, status, "", ""])
                continue

            if allowed == set():
                ws.cell(tag_cell.row, cols["status"], "No mapped PDF sheet")
                ws.cell(tag_cell.row, cols["rating"], "Low")
                ws.cell(tag_cell.row, cols["allowed_pages"], allowed_label)
                ws.cell(tag_cell.row, cols["sheet_filter"], sheet_filter_used)
                ws.cell(tag_cell.row, cols["filter_reason"], filter_reason)
                tag_cell.fill = MISS_FILL
                stats["no_page"] += 1
                result_rows.append([ws_name, tag_cell.row, raw, raw_sheet,
                                    "No mapped PDF sheet", parsed["full_tag"], "", "Low",
                                    0, "", "", "", "", "", "", "", "",
                                    allowed_label, sheet_filter_used, filter_reason, "", ""])
                continue
            if allowed is None:
                stats["all_pages"] += 1

            matches = _choose_matches(parsed, allowed, by_full, by_root_comp,
                                       by_root_only=by_root_only,
                                       by_comp_only=by_comp_only)
            samepage_hit = bool(matches) and matches[0].method == "samepage:root+comp"
            if not matches:
                fallback_hits = _fallback_block_hits(parsed, allowed, block_index)
                if fallback_hits:
                    matches = [_fallback_tag(parsed, hit, len(fallback_hits)) for hit in fallback_hits]
                    stats["fallback"] += 1
                else:
                    ws.cell(tag_cell.row, cols["status"], "Not found in PDF")
                    ws.cell(tag_cell.row, cols["rating"], "Low")
                    ws.cell(tag_cell.row, cols["allowed_pages"], allowed_label)
                    ws.cell(tag_cell.row, cols["sheet_filter"], sheet_filter_used)
                    ws.cell(tag_cell.row, cols["filter_reason"], filter_reason)
                    tag_cell.fill = MISS_FILL
                    stats["not_found"] += 1
                    result_rows.append([ws_name, tag_cell.row, raw, raw_sheet,
                                        "Not found in PDF", parsed["full_tag"], "", "Low",
                                        0, "", "", "", "", "", "", "", "",
                                        allowed_label, sheet_filter_used, filter_reason, "", ""])
                    continue
            elif samepage_hit:
                stats["samepage"] += 1

            best = matches[0]
            conf, rating, extra_evidence = _match_confidence(best, allowed, len(matches))
            if best.method == "block:fallback":
                status = "Matched - block fallback" if len(matches) == 1 else "Matched - block fallback multiple"
            elif best.method == "samepage:root+comp":
                status = "Matched - same-page inference"
            else:
                status = "Matched" if len(matches) == 1 else "Matched - multiple PDF occurrences"
            evidence = best.evidence
            if extra_evidence:
                evidence = f"{evidence}; {extra_evidence}" if evidence else extra_evidence

            tag_cell.fill = _status_fill(status, rating)
            ws.cell(tag_cell.row, cols["status"], status)
            ws.cell(tag_cell.row, cols["confidence"], conf).number_format = "0.00"
            ws.cell(tag_cell.row, cols["rating"], rating)
            ws.cell(tag_cell.row, cols["count"], len(matches))
            ws.cell(tag_cell.row, cols["sheet"], best.sheet_id)
            ws.cell(tag_cell.row, cols["page"], best.page_num)
            ws.cell(tag_cell.row, cols["zone"], best.zone)
            ws.cell(tag_cell.row, cols["method"], best.method)
            ws.cell(tag_cell.row, cols["evidence"], evidence)
            ws.cell(tag_cell.row, cols["text_source"], best.text_source)
            ws.cell(tag_cell.row, cols["source_agreement"], best.source_agreement)
            ws.cell(tag_cell.row, cols["ocr_confidence"],
                    "" if best.ocr_confidence is None else best.ocr_confidence)
            ws.cell(tag_cell.row, cols["allowed_pages"], allowed_label)
            ws.cell(tag_cell.row, cols["sheet_filter"], sheet_filter_used)
            ws.cell(tag_cell.row, cols["filter_reason"], filter_reason)
            ws.cell(tag_cell.row, cols["source_node_ids"], best.source_node_ids)
            ws.cell(tag_cell.row, cols["source_node_count"], best.source_node_count)

            stats["matched"] += 1
            stats[rating[0]] += 1  # H / M / L
            if len(matches) > 1:
                stats["ambiguous"] += 1
            for m in matches:
                key = (m.page_num, m.full_tag, round(m.x0), round(m.y0))
                if key not in matched_keys:
                    matched_keys.add(key)
                    matched_tags.append(m)
            result_rows.append([ws_name, tag_cell.row, raw, raw_sheet, status,
                                parsed["full_tag"], conf, rating, len(matches),
                                best.sheet_id, best.page_num, best.zone,
                                best.method, evidence, best.text_source,
                                best.source_agreement,
                                "" if best.ocr_confidence is None else best.ocr_confidence,
                                allowed_label, sheet_filter_used, filter_reason,
                                best.source_node_ids, best.source_node_count])

    if not found_a_header:
        raise ValueError("Excel file has no recognizable tag column "
                         "(expected one of: " + ", ".join(TAG_HEADERS) + ").")

    return wb, result_rows, stats, matched_tags


def _write_report_sheets(wb, result_rows, stats, all_tags):
    for name in ("Comparison Summary", "Comparison Results",
                 "PDF Candidates Not In Excel", "PDF Page Diagnostics",
                 "Review Queue", "Summary", "Scan Results"):
        if name in wb.sheetnames:
            del wb[name]

    summary = wb.create_sheet("Comparison Summary", 0)
    summary.append(["Metric", "Value"])
    _style_header(summary[1])
    rate = stats["matched"] / stats["reviewed"] if stats["reviewed"] else 0
    cnt = confidence_counts(all_tags)
    rows = [
        ("Rows reviewed", stats["reviewed"]),
        ("Matched rows", stats["matched"]),
        ("Match rate", rate),
        ("High confidence matches", stats["H"]),
        ("Medium confidence matches", stats["M"]),
        ("Low confidence matches", stats["L"]),
        ("Ambiguous matches (multi-occurrence)", stats["ambiguous"]),
        ("Rows without mapped PDF page", stats["no_page"]),
        ("Rows not found in PDF", stats["not_found"]),
        ("Rows skipped (unrecognized)", stats["skipped"]),
        ("Rows matched by block fallback", stats.get("fallback", 0)),
        ("Rows matched by same-page inference", stats.get("samepage", 0)),
        ("Rows searched across all PDF pages", stats["all_pages"]),
        ("PDF candidates total", len(all_tags)),
        ("PDF High confidence candidates", cnt["High"]),
        ("PDF Medium confidence candidates", cnt["Medium"]),
        ("PDF Low confidence candidates", cnt["Low"]),
    ]
    for label, value in rows:
        summary.append([label, value])
    summary["B4"].number_format = "0.00%"
    _format_sheet(summary, [38, 18])

    results = wb.create_sheet("Comparison Results", 1)
    headers = ["Excel Tag", "Excel P&ID", "Status", "Normalized Tag", 
               "Best Confidence", "Rating", "PDF Occurrences"]
    results.append(headers)
    _style_header(results[1])
    for row in result_rows:
        filtered_row = row[2:9]
        results.append(filtered_row)
        status = filtered_row[2]
        rating = filtered_row[5]
        for cell in results[results.max_row]:
            cell.fill = _status_fill(status, rating)
        if isinstance(filtered_row[4], (int, float)):
            results.cell(results.max_row, 5).number_format = "0.00"
    _format_sheet(results, [30, 16, 34, 34, 14, 10, 14])


def _append_pdf_review(output_xlsx, unmatched_tags, profiles):
    wb = load_workbook(output_xlsx)
    ws = wb.create_sheet("PDF Candidates Not In Excel", 2)
    ws.append(["Comparison Status", "Sheet ID", "PDF Page", "Full Tag", "Root", "Component",
               "Class", "Confidence", "Rating", "Zone", "Detection Method", "Evidence",
               "Source Text", "Text Source", "Source Agreement", "Page OCR Mode",
               "OCR Backend", "OCR Reason", "OCR Confidence", "Source Node IDs",
               "Source Node Count"])
    _style_header(ws[1])
    for t in sorted(unmatched_tags, key=lambda x: (x.page_num, x.sheet_id, x.zone, x.tag_class, -x.confidence, x.full_tag)):
        ws.append(["Missing in Excel", t.sheet_id, t.page_num, t.full_tag, t.root, t.component,
                   t.tag_class, round(t.confidence, 3),
                   t.rating, t.zone, t.method, t.evidence, t.source_text,
                   t.text_source, t.source_agreement, t.page_ocr_mode, t.ocr_backend,
                   t.ocr_reason, "" if t.ocr_confidence is None else t.ocr_confidence,
                   t.source_node_ids, t.source_node_count])
        for cell in ws[ws.max_row]:
            cell.fill = LOW_FILL
        ws.cell(ws.max_row, 8).number_format = "0.00"
    _format_sheet(ws, [18, 14, 10, 36, 24, 18, 14, 12, 10, 22, 20, 52, 60,
                       14, 18, 16, 16, 40, 14, 44, 14])

    _save(wb, output_xlsx)


def _passes_annotation_threshold(tag, threshold):
    return CONFIDENCE_RANK.get(tag.rating, 0) >= CONFIDENCE_RANK.get(threshold.title(), 2)


def _match_annotation_color(tag):
    return YELLOW_HIGHLIGHT if tag.rating == "High" else RED_HIGHLIGHT


def _is_precise_annotation_rect(rect):
    """Reject broad union/block rectangles; keep tight text-span boxes."""
    width = abs(rect.x1 - rect.x0)
    height = abs(rect.y1 - rect.y0)
    area = width * height
    # Horizontal labels should stay close to a line of text. Tall/narrow boxes
    # are allowed for rotated text, but broad area rectangles are rejected.
    if width <= 0 or height <= 0:
        return False
    if width > 180 and height > 35:
        return False
    if height > 180 and width > 35:
        return False
    return width <= 220 and height <= 180 and area <= 9000


def _rect_overlap_ratio(a, b):
    inter = fitz.Rect(a)
    inter.intersect(b)
    if inter.is_empty:
        return 0.0
    ia = max(0, inter.width) * max(0, inter.height)
    aa = max(1.0, abs(a.width * a.height))
    ba = max(1.0, abs(b.width * b.height))
    return ia / min(aa, ba)


def _display_rects_for_annotation(tag):
    rects = []
    for raw in getattr(tag, "source_rects", ()) or ():
        if len(raw) != 4:
            continue
        rect = fitz.Rect(*raw)
        rect.normalize()
        if _is_precise_annotation_rect(rect):
            rects.append(rect)
    if rects:
        return rects

    # Block fallback rectangles are whole PDF text blocks and create the large
    # highlights the team does not want. Only fall back to the tag bbox for
    # native scanner detections whose bbox is already text-sized.
    if tag.method == "block:fallback":
        return []
    rect = fitz.Rect(tag.x0, tag.y0, tag.x1, tag.y1)
    rect.normalize()
    return [rect] if _is_precise_annotation_rect(rect) else []


def _annotate_pdf(pdf_path, matched_tags, unmatched_tags=None, annotation_confidence="medium"):
    """Highlight precise detected text spans, not broad combined blocks."""
    base, ext = os.path.splitext(pdf_path)
    output = base + "_annotated" + ext
    pages_touched = set()
    with fitz.open(pdf_path) as doc:
        # Source PDFs may already contain large legacy markups from Bluebeam or
        # previous tool runs. The annotated output should show only this run's
        # precise scanner/comparer highlights.
        for page in doc:
            for annot in list(page.annots() or []):
                page.delete_annot(annot)

        candidates = []
        for t in matched_tags:
            if not _passes_annotation_threshold(t, annotation_confidence):
                continue
            color = _match_annotation_color(t)
            priority = 3 if color == YELLOW_HIGHLIGHT else 2
            for rect in _display_rects_for_annotation(t):
                candidates.append((priority, t.page_num, rect, color))
        for t in unmatched_tags or []:
            if not _passes_annotation_threshold(t, annotation_confidence):
                continue
            for rect in _display_rects_for_annotation(t):
                candidates.append((1, t.page_num, rect, RED_HIGHLIGHT))

        # Remove stacked/double/triple highlights. If the same text area is
        # found through multiple paths, keep the highest-priority color:
        # green matched > yellow matched > red PDF-only.
        accepted = []
        for priority, page_num, rect, color in sorted(candidates, key=lambda x: -x[0]):
            if any(page_num == ap and _rect_overlap_ratio(rect, ar) > 0.20
                   for _, ap, ar, _ in accepted):
                continue
            accepted.append((priority, page_num, rect, color))

        for _, page_num, display_rect, color in accepted:
            page = doc[page_num - 1]
            native_rect = display_rect * page.derotation_matrix
            native_rect.normalize()
            try:
                annot = page.add_highlight_annot(native_rect)
                annot.set_colors(stroke=color)
                annot.set_opacity(0.35)
            except Exception:
                annot = page.add_rect_annot(native_rect)
                annot.set_colors(stroke=color, fill=color)
                annot.set_opacity(0.20)
                annot.set_border(width=0.5)
            annot.update()
            pages_touched.add(page_num)
        doc.save(output, incremental=False, encryption=fitz.PDF_ENCRYPT_NONE)
    return output, len(pages_touched)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def run_comparison(pdf_path, xlsx_path, mode="all", annotate_pdf=True,
                   annotate_unmatched_pdf=True, annotation_confidence="medium",
                   ocr_mode="auto", ocr_dpi=300, debug_text=False):
    pdf_path = validate_input_file(pdf_path, {".pdf"})
    xlsx_path = validate_input_file(xlsx_path, {".xlsx"})
    validate_pdf_content(pdf_path)
    validate_xlsx(xlsx_path)

    with fitz.open(pdf_path) as doc:
        sheet_map, page_to_sheet = build_sheet_map(doc)
        block_index = _build_block_fallback_index(doc, page_to_sheet)
        mapped_page_count = sum(len(v) if isinstance(v, (list, tuple, set)) else 1
                                for v in sheet_map.values())
        raw_tags, profiles = scan_open_document(
            doc,
            page_to_sheet=page_to_sheet,
            mode=mode,
            ocr_mode=ocr_mode,
            ocr_dpi=ocr_dpi,
            debug_text=debug_text,
        )
    tags = dedupe(raw_tags, by_location=True)

    wb, result_rows, stats, matched_tags = process_excel(
        xlsx_path,
        sheet_map,
        tags,
        block_index=block_index,
        mode=mode,
    )
    _write_report_sheets(wb, result_rows, stats, tags)

    base, ext = os.path.splitext(xlsx_path)
    output_xlsx = _save(wb, base + "_annotated" + ext)
    matched_keys = {(m.page_num, m.full_tag, round(m.x0), round(m.y0))
                    for m in matched_tags}
    unmatched_tags = [t for t in tags
                      if (t.page_num, t.full_tag, round(t.x0), round(t.y0))
                      not in matched_keys]
    _append_pdf_review(output_xlsx, unmatched_tags, profiles)

    output_pdf = None
    pages_annotated = 0
    unmatched_for_annotation = unmatched_tags if annotate_unmatched_pdf else []
    if annotate_pdf and (matched_tags or unmatched_for_annotation):
        output_pdf, pages_annotated = _annotate_pdf(
            pdf_path,
            matched_tags,
            unmatched_for_annotation,
            annotation_confidence=annotation_confidence,
        )

    return {
        "pdf_path": pdf_path,
        "xlsx_path": xlsx_path,
        "output_xlsx": output_xlsx,
        "output_pdf": output_pdf,
        "pages_mapped": mapped_page_count,
        "pages_annotated": pages_annotated,
        "rows_reviewed": stats["reviewed"],
        "matched": stats["matched"],
        "matched_high": stats["H"],
        "matched_medium": stats["M"],
        "matched_low": stats["L"],
        "not_found": stats["not_found"],
        "no_page": stats["no_page"],
        "skipped": stats["skipped"],
        "ambiguous": stats["ambiguous"],
        "ocr_mode_requested": ocr_mode,
        "ocr_dpi": ocr_dpi,
        "annotate_unmatched_pdf": annotate_unmatched_pdf,
        "annotation_confidence": annotation_confidence,
        **summarize_profiles(profiles),
    }


def _cli():
    parser = argparse.ArgumentParser(description="Compare an Excel tag list against a P&ID PDF.")
    parser.add_argument("pdf")
    parser.add_argument("xlsx")
    parser.add_argument("--mode", default="all",
                        choices=["instrument", "equipment", "line", "all"])
    parser.add_argument("--no-annotate-pdf", action="store_true")
    parser.add_argument("--annotate-unmatched-pdf", dest="annotate_unmatched_pdf",
                        action="store_true", default=True,
                        help="Also annotate PDF-only candidates in red (default: on)")
    parser.add_argument("--no-annotate-unmatched-pdf", dest="annotate_unmatched_pdf",
                        action="store_false",
                        help="Disable red annotations on the PDF")
    parser.add_argument("--annotation-confidence", default="medium",
                        choices=["high", "medium", "all"],
                        help="Minimum candidate rating to annotate (default: medium)")
    parser.add_argument("--ocr", default="auto", choices=["auto", "always", "off"],
                        help="OCR routing mode (default: auto)")
    parser.add_argument("--ocr-dpi", type=int, default=300,
                        help="DPI for PyMuPDF OCR when OCR runs (default: 300)")
    parser.add_argument("--debug-text", action="store_true")
    args = parser.parse_args()
    result = run_comparison(args.pdf, args.xlsx, mode=args.mode,
                            annotate_pdf=not args.no_annotate_pdf,
                            annotate_unmatched_pdf=args.annotate_unmatched_pdf,
                            annotation_confidence=args.annotation_confidence,
                            ocr_mode=args.ocr,
                            ocr_dpi=args.ocr_dpi,
                            debug_text=args.debug_text)
    print(f"Comparison complete")
    print(f"  Excel -> {result['output_xlsx']}")
    if result['output_pdf']:
        print(f"  PDF   -> {result['output_pdf']}")
    print(f"  reviewed: {result['rows_reviewed']}  matched: {result['matched']} "
          f"(H={result['matched_high']} M={result['matched_medium']} L={result['matched_low']})")
    print(f"  not found: {result['not_found']}  no page: {result['no_page']} "
          f"skipped: {result['skipped']}  ambiguous: {result['ambiguous']}")
    print(f"  ocr: {result['ocr_mode_requested']} dpi={result['ocr_dpi']}  "
          f"annotation confidence: {result['annotation_confidence']}")


if __name__ == "__main__":
    _cli()
