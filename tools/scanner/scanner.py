"""Scanner: read a P&ID PDF, extract tags, write a multi-sheet Excel report.

Usage (CLI):
    python scanner/scanner.py <pdf> [--mode instrument|equipment|line|all]

Usage (library):
    from scanner.scanner import run_scan
    result = run_scan("mydrawing.pdf", mode="all")
"""

import argparse
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


TOOL_DIR = os.path.abspath(os.path.dirname(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(TOOL_DIR, os.pardir))
for path in (TOOL_DIR, PROJECT_ROOT):
    if path not in sys.path:
        sys.path.insert(0, path)

from .pid_common import (
    validate_input_file,
    validate_pdf_content,
)
from .pid_pdf_common import (
    confidence_counts,
    dedupe,
    scan_document,
    summarize_profiles,
)


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HIGH_FILL = PatternFill("solid", fgColor="FFF2CC")   # light yellow
MED_FILL = PatternFill("solid", fgColor="F4CCCC")    # light red
LOW_FILL = PatternFill("solid", fgColor="F4CCCC")    # light red
DETECTION_REVIEW_FILL = MED_FILL

SCAN_HEADERS = [
    "Sheet ID", "PDF Page", "Full Tag", "Root", "Component",
    "Confidence", "Rating", "Source Text",
]
SCAN_WIDTHS = [14, 10, 36, 24, 18, 12, 10, 60]

REVIEW_HEADERS = [
    "Sheet ID", "PDF Page", "Full Tag", "Root",
    "Confidence", "Rating",
]
REVIEW_WIDTHS = [14, 10, 36, 24, 12, 10]


def _style_header(row):
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _format_sheet(sheet, widths):
    sheet.freeze_panes = "A2"
    if sheet.max_row > 1:
        sheet.auto_filter.ref = sheet.dimensions
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width


def _rating_fill(rating):
    return HIGH_FILL if rating == "High" else DETECTION_REVIEW_FILL


def _save_workbook(wb, preferred_path):
    """Save with a numeric suffix fallback if the target file is locked (Excel open)."""
    try:
        wb.save(preferred_path)
        return preferred_path
    except PermissionError:
        base, ext = os.path.splitext(preferred_path)
        for i in range(1, 100):
            fallback = f"{base}_{i}{ext}"
            try:
                wb.save(fallback)
                return fallback
            except PermissionError:
                continue
        raise


def _tag_row(t):
    return [
        t.sheet_id, t.page_num, t.full_tag, t.root, t.component,
        round(t.confidence, 3), t.rating, t.source_text,
    ]


def _review_row(t):
    return [
        t.sheet_id, t.page_num, t.full_tag, t.root,
        round(t.confidence, 3), t.rating,
    ]


def _report_sort_key(t):
    zone_rank = {"drawing": 0, "title-block/right": 1, "title-block/bottom": 2, "edge": 3}
    class_rank = {"instrument": 0, "equipment": 1, "line": 2, "spec": 3}
    return (
        t.page_num,
        t.sheet_id,
        zone_rank.get(t.zone, 9),
        class_rank.get(t.tag_class, 9),
        -t.confidence,
        t.full_tag,
    )


def write_report(tags, profiles, pdf_path, mode="all", raw_count=None, ocr_mode="auto", ocr_dpi=300, debug_text=False):
    tags = dedupe(tags)
    counts = confidence_counts(tags)
    profile_summary = summarize_profiles(profiles)

    wb = Workbook()

    # --- Scan Results ---
    sh = wb.active
    sh.title = "Scan Results"
    sh.append(SCAN_HEADERS)
    _style_header(sh[1])
    for t in sorted(tags, key=_report_sort_key):
        sh.append(_tag_row(t))
        for cell in sh[sh.max_row]:
            cell.fill = _rating_fill(t.rating)
        sh.cell(sh.max_row, 6).number_format = "0.00"
    _format_sheet(sh, SCAN_WIDTHS)

    # --- Summary ---
    summary = wb.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    _style_header(summary[1])
    rows = [
        ("Scan mode", mode),
        ("OCR mode", ocr_mode),
        ("OCR DPI", ocr_dpi),
        ("Unique tags reported", len(tags)),
        ("Raw occurrences", raw_count if raw_count is not None else len(tags)),
        ("Duplicates collapsed", (raw_count or len(tags)) - len(tags)),
        ("Unique full tags", len({t.full_tag for t in tags})),
        ("Sheets with candidates", len({t.sheet_id for t in tags})),
        ("High confidence", counts["High"]),
        ("Medium confidence", counts["Medium"]),
        ("Low confidence", counts["Low"]),
        ("Instruments", sum(1 for t in tags if t.tag_class == "instrument")),
        ("Equipment", sum(1 for t in tags if t.tag_class == "equipment")),
        ("Line tags", sum(1 for t in tags if t.tag_class == "line")),
        ("Pages scanned", profile_summary["pages"]),
        ("Pages needing OCR", profile_summary["ocr_needed"]),
        ("Pages OCR used", profile_summary["ocr_used"]),
        ("OCR warnings", profile_summary["warnings"]),
        ("Avg words per page", profile_summary["avg_words"]),
        ("Native spans", profile_summary["native_spans"]),
        ("OCR spans", profile_summary["ocr_spans"]),
        ("Graph groups", profile_summary["graph_groups"]),
    ]
    for label, value in rows:
        summary.append([label, value])
    _format_sheet(summary, [32, 18])

    # --- Review Queue (non-High rows) ---
    review = wb.create_sheet("Review Queue")
    review.append(REVIEW_HEADERS)
    _style_header(review[1])
    for t in sorted(tags, key=_report_sort_key):
        if t.rating == "High":
            continue
        review.append(_review_row(t))
        for cell in review[review.max_row]:
            cell.fill = _rating_fill(t.rating)
        review.cell(review.max_row, 5).number_format = "0.00"
    _format_sheet(review, REVIEW_WIDTHS)

    if debug_text:
        tokens = wb.create_sheet("Text Tokens")
        token_headers = ["Sheet ID", "PDF Page", "Text Source", "Span ID", "Node ID",
                         "X0", "Y0", "X1", "Y1", "Font Size", "Text", "Normalized Text"]
        tokens.append(token_headers)
        _style_header(tokens[1])
        for prof in profiles:
            for row in prof.text_tokens:
                tokens.append([
                    row["sheet_id"], row["page_num"], row["source"], row["span_id"],
                    row["node_id"], row["x0"], row["y0"], row["x1"], row["y1"],
                    row["size"], row["text"], row["norm"],
                ])
        _format_sheet(tokens, [14, 10, 14, 12, 24, 10, 10, 10, 10, 10, 60, 60])

        groups = wb.create_sheet("Text Groups")
        group_headers = ["Sheet ID", "PDF Page", "Text Source", "Group ID", "Node IDs",
                         "X0", "Y0", "X1", "Y1", "Text", "Normalized Text"]
        groups.append(group_headers)
        _style_header(groups[1])
        for prof in profiles:
            for row in prof.text_groups:
                groups.append([
                    row["sheet_id"], row["page_num"], row["source"], row["group_id"],
                    row["node_ids"], row["x0"], row["y0"], row["x1"], row["y1"],
                    row["text"], row["norm"],
                ])
        _format_sheet(groups, [14, 10, 14, 16, 44, 10, 10, 10, 10, 60, 60])

        rejected = wb.create_sheet("Rejected Text")
        rejected.append(["Sheet ID", "PDF Page", "Text Source", "Rejected Text"])
        _style_header(rejected[1])
        for prof in profiles:
            for row in prof.rejected_text:
                rejected.append([row["sheet_id"], row["page_num"], row["source"], row["text"]])
        _format_sheet(rejected, [14, 10, 14, 60])

        graph_candidates = wb.create_sheet("Graph Candidates")
        graph_candidates.append(SCAN_HEADERS)
        _style_header(graph_candidates[1])
        for t in sorted(tags, key=_report_sort_key):
            graph_candidates.append(_tag_row(t))
            for cell in graph_candidates[graph_candidates.max_row]:
                cell.fill = _rating_fill(t.rating)
            graph_candidates.cell(graph_candidates.max_row, 6).number_format = "0.00"
        _format_sheet(graph_candidates, SCAN_WIDTHS)

    base, _ = os.path.splitext(pdf_path)
    return _save_workbook(wb, base + "_scan_report.xlsx")


def run_scan(pdf_path, mode="all", ocr_mode="auto", ocr_dpi=300, debug_text=False):
    pdf_path = validate_input_file(pdf_path, {".pdf"})
    validate_pdf_content(pdf_path)
    raw_tags, profiles = scan_document(
        pdf_path,
        mode=mode,
        ocr_mode=ocr_mode,
        ocr_dpi=ocr_dpi,
        debug_text=debug_text,
    )
    tags = dedupe(raw_tags)
    output = write_report(
        tags,
        profiles,
        pdf_path,
        mode=mode,
        raw_count=len(raw_tags),
        ocr_mode=ocr_mode,
        ocr_dpi=ocr_dpi,
        debug_text=debug_text,
    )
    counts = confidence_counts(tags)
    return {
        "pdf_path": pdf_path,
        "output_excel": output,
        "mode": mode,
        "ocr_mode_requested": ocr_mode,
        "ocr_dpi": ocr_dpi,
        "total_tags": len(tags),
        "raw_occurrences": len(raw_tags),
        "unique_full_tags": len({t.full_tag for t in tags}),
        "high": counts["High"],
        "medium": counts["Medium"],
        "low": counts["Low"],
        **summarize_profiles(profiles),
    }


def _cli():
    parser = argparse.ArgumentParser(description="Scan a P&ID PDF for tags.")
    parser.add_argument("pdf", help="Path to input PDF")
    parser.add_argument("--mode", default="all",
                        choices=["instrument", "equipment", "line", "all"],
                        help="Which class(es) of tag to extract (default: all)")
    parser.add_argument("--ocr", default="auto", choices=["auto", "always", "off"],
                        help="OCR routing mode (default: auto)")
    parser.add_argument("--ocr-dpi", type=int, default=300,
                        help="DPI for PyMuPDF OCR when OCR runs (default: 300)")
    parser.add_argument("--debug-text", action="store_true",
                        help="Add text token, group, rejected text, and graph candidate sheets")
    args = parser.parse_args()
    result = run_scan(args.pdf, mode=args.mode, ocr_mode=args.ocr,
                      ocr_dpi=args.ocr_dpi, debug_text=args.debug_text)
    print(f"Scan complete -> {result['output_excel']}")
    print(f"  mode: {result['mode']}")
    print(f"  ocr: {result['ocr_mode_requested']} dpi={result['ocr_dpi']}")
    print(f"  tags: {result['total_tags']} "
          f"(H={result['high']} M={result['medium']} L={result['low']})")
    print(f"  pages: {result['pages']} (OCR used on {result['ocr_used']})")


if __name__ == "__main__":
    _cli()
